"""报告导出服务."""

import json
from collections.abc import Callable
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.models.report import Report
from app.models.user import User
from app.services.audit_service import log_action
from app.storage import BaseStorageClient


class ExportFormatError(Exception):
    """不支持的导出格式."""

    pass


# 容器内文泉驿微米黑字体路径；若存在则使用 Unicode 字体渲染中文，否则安全降级
_CJK_FONT_PATH = Path("/usr/share/fonts/truetype/wqy/wqy-microhei.ttc")
_CJK_FONT_FAMILY = "WQYMicroHei"


def _render_markdown(report: Report) -> bytes:
    """将报告内容渲染为 Markdown 字节."""
    content = report.content or {}
    lines: list[str] = []
    lines.append(f"# {content.get('title', report.title)}")
    lines.append("")
    lines.append(f"- 报告类型：{report.report_type}")
    lines.append(f"- 生成时间：{report.created_at.isoformat() if report.created_at else ''}")
    lines.append(f"- 导出时间：{datetime.now(UTC).isoformat()}")
    lines.append("")

    summary = content.get("summary") or report.summary
    if summary:
        lines.append("## 摘要")
        lines.append(summary)
        lines.append("")

    sections = content.get("sections") or []
    if sections:
        lines.append("## 指标")
        lines.append("")
        lines.append("| 指标 | 数值 |")
        lines.append("|------|------|")
        for section in sections:
            name = section.get("name", "")
            value = section.get("value", "")
            lines.append(f"| {name} | {value} |")
        lines.append("")

    return "\n".join(lines).encode("utf-8")


def _render_json(report: Report) -> bytes:
    """将报告完整内容渲染为 JSON 字节."""
    payload = {
        "id": report.id,
        "title": report.title,
        "report_type": report.report_type,
        "parameters": report.parameters,
        "content": report.content,
        "summary": report.summary,
        "created_at": report.created_at.isoformat() if report.created_at else None,
        "exported_at": datetime.now(UTC).isoformat(),
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def _safe_pdf_text(text: str) -> str:
    """将文本编码为 PDF 核心字体支持的 latin-1 字符集."""
    return text.encode("latin-1", "replace").decode("latin-1")


def _has_cjk_font() -> bool:
    """是否已安装可用于 PDF 的中文 Unicode 字体."""
    return _CJK_FONT_PATH.exists()


def _pdf_text(text: str) -> str:
    """根据字体可用性返回可直接写入 PDF 的文本."""
    if _has_cjk_font():
        return text
    return _safe_pdf_text(text)


def _render_pdf(report: Report) -> bytes:
    """将报告内容渲染为 PDF 字节."""
    content = report.content or {}
    title = content.get("title", report.title)
    summary = content.get("summary") or report.summary or ""
    sections = content.get("sections") or []

    pdf = FPDF()
    pdf.add_page()

    use_cjk = _has_cjk_font()
    if use_cjk:
        pdf.add_font(_CJK_FONT_FAMILY, "", str(_CJK_FONT_PATH))
        pdf.set_font(_CJK_FONT_FAMILY, "", 16)
    else:
        pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, _pdf_text(title), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    if use_cjk:
        pdf.set_font(_CJK_FONT_FAMILY, "", 12)
    else:
        pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 8, _pdf_text(f"报告类型：{report.report_type}"), new_x="LMARGIN", new_y="NEXT")
    pdf.cell(
        0,
        8,
        _pdf_text(f"生成时间：{report.created_at.isoformat() if report.created_at else ''}"),
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.cell(
        0,
        8,
        _pdf_text(f"导出时间：{datetime.now(UTC).isoformat()}"),
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.ln(5)

    if summary:
        if use_cjk:
            pdf.set_font(_CJK_FONT_FAMILY, "", 14)
        else:
            pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, _pdf_text("摘要"), new_x="LMARGIN", new_y="NEXT")
        if use_cjk:
            pdf.set_font(_CJK_FONT_FAMILY, "", 12)
        else:
            pdf.set_font("Helvetica", "", 12)
        pdf.multi_cell(0, 8, _pdf_text(summary))
        pdf.ln(5)

    if sections:
        if use_cjk:
            pdf.set_font(_CJK_FONT_FAMILY, "", 14)
        else:
            pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, _pdf_text("指标"), new_x="LMARGIN", new_y="NEXT")
        if use_cjk:
            pdf.set_font(_CJK_FONT_FAMILY, "", 12)
        else:
            pdf.set_font("Helvetica", "B", 12)
        pdf.cell(90, 8, _pdf_text("指标"), border=1)
        pdf.cell(90, 8, _pdf_text("数值"), border=1, new_x="LMARGIN", new_y="NEXT")
        if use_cjk:
            pdf.set_font(_CJK_FONT_FAMILY, "", 12)
        else:
            pdf.set_font("Helvetica", "", 12)
        for section in sections:
            name = section.get("name", "")
            value = section.get("value", "")
            pdf.cell(90, 8, _pdf_text(str(name)), border=1)
            pdf.cell(90, 8, _pdf_text(str(value)), border=1, new_x="LMARGIN", new_y="NEXT")

    return bytes(pdf.output())


def _render_excel(report: Report) -> bytes:
    """将报告内容渲染为 Excel 字节."""
    content = report.content or {}
    title = content.get("title", report.title)
    summary = content.get("summary") or report.summary or ""
    sections = content.get("sections") or []

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"

    worksheet["A1"] = title
    worksheet["A1"].font = Font(bold=True, size=16)

    worksheet["A2"] = "Report Type"
    worksheet["B2"] = report.report_type
    worksheet["A3"] = "Created At"
    worksheet["B3"] = report.created_at.isoformat() if report.created_at else ""
    worksheet["A4"] = "Exported At"
    worksheet["B4"] = datetime.now(UTC).isoformat()

    row = 6
    if summary:
        worksheet.cell(row=row, column=1, value="Summary").font = Font(bold=True, size=14)
        row += 1
        worksheet.cell(row=row, column=1, value=summary)
        row += 2

    if sections:
        worksheet.cell(row=row, column=1, value="Metrics").font = Font(bold=True, size=14)
        row += 1
        worksheet.cell(row=row, column=1, value="Metric").font = Font(bold=True)
        worksheet.cell(row=row, column=2, value="Value").font = Font(bold=True)
        row += 1
        for section in sections:
            worksheet.cell(row=row, column=1, value=section.get("name", ""))
            worksheet.cell(row=row, column=2, value=section.get("value", ""))
            row += 1

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class ExportRendererRegistry:
    """报告导出渲染器注册表.

    新增导出格式只需调用 `register(fmt, renderer, content_type, ext)`，
    无需修改 export_report 主逻辑，便于后续扩展 Word/CSV/图片等格式。
    """

    Entry = tuple[Callable[[Report], bytes], str, str]

    _renderers: dict[str, Entry] = {}

    @classmethod
    def register(
        cls,
        fmt: str,
        content_type: str,
        ext: str,
    ) -> Callable[[Callable[[Report], bytes]], Callable[[Report], bytes]]:
        """注册导出格式渲染器（用作装饰器）."""

        def decorator(func: Callable[[Report], bytes]) -> Callable[[Report], bytes]:
            cls._renderers[fmt] = (func, content_type, ext)
            return func

        return decorator

    @classmethod
    def get_renderer(cls, fmt: str) -> tuple[Callable[[Report], bytes], str, str]:
        """获取格式对应的渲染函数、Content-Type 与扩展名."""
        entry = cls._renderers.get(fmt)
        if entry is None:
            raise ExportFormatError(f"不支持的导出格式: {fmt}")
        return entry

    @classmethod
    def list_formats(cls) -> list[str]:
        """返回所有已注册导出格式."""
        return list(cls._renderers.keys())


@ExportRendererRegistry.register("markdown", "text/markdown; charset=utf-8", "md")
def _render_markdown_registered(report: Report) -> bytes:
    return _render_markdown(report)


@ExportRendererRegistry.register("json", "application/json; charset=utf-8", "json")
def _render_json_registered(report: Report) -> bytes:
    return _render_json(report)


@ExportRendererRegistry.register("pdf", "application/pdf", "pdf")
def _render_pdf_registered(report: Report) -> bytes:
    return _render_pdf(report)


@ExportRendererRegistry.register(
    "xlsx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "xlsx",
)
def _render_excel_registered(report: Report) -> bytes:
    return _render_excel(report)


def export_report(
    db: Session,
    report: Report,
    storage: BaseStorageClient,
    user: User | None,
    fmt: str = "markdown",
) -> str:
    """导出报告到对象存储.

    Args:
        db: 数据库会话。
        report: 已生成的报告。
        storage: 对象存储客户端（本地或 MinIO 均可）。
        user: 当前用户，用于审计日志。
        fmt: 导出格式，支持 markdown/json/pdf/xlsx。

    Returns:
        导出的文件访问 URL。

    Raises:
        ExportFormatError: 格式不支持。
        StorageClientError: 上传失败。
    """
    renderer, content_type, ext = ExportRendererRegistry.get_renderer(fmt)

    data = renderer(report)
    key = f"reports/{report.tenant_id}/{report.id}/report.{ext}"

    url = storage.upload_bytes(
        key=key,
        data=data,
        content_type=content_type,
        metadata={"report_id": report.id, "report_type": report.report_type},
    )

    report.content_url = url

    log_action(
        db=db,
        action="report.export",
        resource=f"report://{report.id}",
        user=user,
        result="success",
        reason=f"format={fmt};url={url}",
    )

    return url
