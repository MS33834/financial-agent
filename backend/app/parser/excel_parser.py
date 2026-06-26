"""Excel 财务数据解析器."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from app.parser.base import BaseDocumentParser, ParserRegistry
from app.parser.utils import extract_period, extract_year

# 单文件最大解析行数，防止恶意大文件耗尽内存
MAX_EXCEL_ROWS = 100_000


@ParserRegistry.register
class ExcelParser(BaseDocumentParser):
    """基于 openpyxl 的 Excel 解析器.

    默认读取第一个工作表，第一行作为表头；
    列名支持中英文别名，由 financial_import_service 统一映射。
    """

    supported_extensions = {"xlsx", "xls"}

    def parse(self, content: bytes, filename: str) -> dict[str, Any]:
        """解析 Excel 内容."""
        try:
            import openpyxl
        except ImportError as exc:
            raise RuntimeError("Excel 解析需要 openpyxl，请安装：pip install openpyxl") from exc

        workbook = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
        sheet = workbook.active
        if sheet is None:
            return self._empty_result(filename)

        # 流式限制行数，防止恶意大文件耗尽内存。
        # 不使用 iter_rows(max_row=...) —— 该参数会对空表填充空行，导致误判。
        # 此处惰性迭代，达到上限即中止并拒绝，避免一次性载入超大文件。
        rows: list[tuple[Any, ...]] = []
        row_limit = MAX_EXCEL_ROWS + 1  # 表头 + 数据上限，超出即拒绝
        for row in sheet.iter_rows(values_only=True):
            rows.append(row)
            if len(rows) > row_limit:
                raise ValueError(
                    f"Excel 数据行数超过上限 {MAX_EXCEL_ROWS}，请拆分文件后重试"
                )
        if len(rows) < 2:
            return self._empty_result(filename)

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        records: list[dict[str, Any]] = []
        for row in rows[1:]:
            record = {header: value for header, value in zip(headers, row, strict=False) if header}
            if any(v is not None and str(v).strip() != "" for v in record.values()):
                records.append(record)

        return {
            "format": "excel",
            "filename": filename,
            "extension": filename.split(".")[-1].lower() if "." in filename else "",
            "sheet_name": sheet.title,
            "detected_year": extract_year(filename),
            "detected_period": extract_period(filename),
            "records": records,
            "tables": [records],
            "row_count": len(records),
            "confidence": 0.95 if records else 0.5,
        }

    def _empty_result(self, filename: str) -> dict[str, Any]:
        return {
            "format": "excel",
            "filename": filename,
            "extension": filename.split(".")[-1].lower() if "." in filename else "",
            "detected_year": extract_year(filename),
            "detected_period": extract_period(filename),
            "records": [],
            "tables": [],
            "row_count": 0,
            "confidence": 0.5,
        }
